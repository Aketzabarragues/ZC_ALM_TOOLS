"""
Infrastructure Layer - Base Parser
===================================
Clase base con lógica genérica para extraer tablas de Excel.
"""

import logging
from typing import cast

import pandas as pd
import openpyxl
from openpyxl.utils import cell as utils_cell
from openpyxl.worksheet.worksheet import Worksheet


class ExcelParsingError(Exception):
    """Excepción base para errores de parsing de Excel."""
    def __init__(self, message: str) -> None:
        super().__init__(message)


class BaseParser:
    """
    Parser base con lógica genérica para localizar y extraer tablas de Excel.

    Utiliza los metadatos de tablas definidas en Excel (ListObject) para
    extraer datos de forma robusta sin hardcodear rangos.
    """

    def __init__(self) -> None:
        self._logger: logging.Logger = logging.getLogger(
            f"{__name__}.{self.__class__.__name__}"
        )

    def _leer_tabla(
        self,
        ruta_excel: str,
        sheet_name: str,
        table_name: str,
        columnas_numericas: list[str],
    ) -> pd.DataFrame:
        """
        Lógica genérica robusta para localizar y extraer una tabla como DataFrame.

        Args:
            ruta_excel: Ruta al archivo Excel.
            sheet_name: Nombre de la hoja que contiene la tabla.
            table_name: Nombre de la tabla (ListObject) a extraer.
            columnas_numericas: Lista de nombres de columnas a cast a int.

        Returns:
            DataFrame con los datos de la tabla.

        Raises:
            ExcelParsingError: Si la tabla o hoja no se encuentra.
        """
        self._logger.info(
            f"Extrayendo tabla '{table_name}' de la hoja '{sheet_name}'..."
        )

        try:
            workbook = openpyxl.load_workbook(filename=ruta_excel, data_only=True)

            if sheet_name not in workbook.sheetnames:
                raise ExcelParsingError(f"Hoja '{sheet_name}' no encontrada.")

            ws: Worksheet = cast(Worksheet, workbook[sheet_name])

            # Buscar tabla por nombre en los metadatos de Excel
            tabla = next(
                (t for t in ws.tables.values() if t.name == table_name),
                None
            )

            if not tabla:
                raise ExcelParsingError(
                    f"Tabla '{table_name}' no encontrada en metadatos."
                )

            # Obtener boundaries del rango
            raw_bounds = utils_cell.range_boundaries(tabla.ref)
            if not raw_bounds:
                raise ExcelParsingError("No se pudieron leer los boundaries.")

            min_col, min_row, max_col, max_row = cast(
                tuple[int, int, int, int], raw_bounds
            )

            # Leer usando iter_rows de openpyxl (una sola lectura en disco)
            data = [
                [cell.value for cell in row]
                for row in ws.iter_rows(
                    min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row
                )
            ]

            if not data:
                self._logger.warning(f"La hoja '{sheet_name}' está vacía.")
                return pd.DataFrame()

            # Si solo tiene cabeceras (sin datos), devolver DataFrame con columnas correctas
            if len(data) < 2:
                cabeceras = data[0] if data else []
                self._logger.warning(f"La tabla '{table_name}' solo contiene cabeceras.")
                return pd.DataFrame(columns=cabeceras)

            # Crear DataFrame: primera fila cabeceras, resto datos
            cabeceras = data[0]
            filas = data[1:]
            df: pd.DataFrame = pd.DataFrame(filas, columns=cabeceras)

            # Eliminar filas sin UID (vacías o encabezado)
            df = df.dropna(subset=["UID"])

            # Cast columnas numéricas
            for col in columnas_numericas:
                if col in df.columns:
                    df[col] = df[col].fillna(0).astype(int)

            # Convertir resto a string puro y limpiar NaNs
            columnas_texto: list[str] = [
                col for col in df.columns if col not in columnas_numericas
            ]
            for col in columnas_texto:
                df[col] = df[col].fillna("").astype(str)

            self._logger.info(f"Tabla '{table_name}' extraída: {len(df)} filas.")
            return df

        except ExcelParsingError:
            raise
        except Exception as e:
            raise ExcelParsingError(
                f"Error al leer '{table_name}': {e}"
            ) from e